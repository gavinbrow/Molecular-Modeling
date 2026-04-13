#!/usr/bin/env python3
"""
orca_run.py — Run ORCA jobs from INP → MID → OUT, then automatically
extract Gibbs free energy, timings, and errors into an Excel report.

Folder structure (both MID and OUT use matching date folders):

    MID/
      Apr. 02 2026 10.37am/
        water/   ← ORCA working directory
        water2/

    OUT/
      Apr. 02 2026 10.37am/
        water/
          water.out
          water.xyz
        water2/
          water2.out
        orca_report.xlsx

Usage:
    orca_run.py jobname [jobname2 ...]
    orca_run.py jobname.inp other.inp
    orca_run.py "C:\\full\\path\\to\\file1.inp"
    orca_run.py -all              Run every .inp in the INP folder

Requires: openpyxl  (pip install openpyxl)
"""

import os
import re
import sys
import time
import shutil
from pathlib import Path
from datetime import datetime
import subprocess

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

ROOT = Path(__file__).resolve().parent

# ===================== USER SETTING =====================
ORCA_EXE = Path(r"C:\ORCA_6.1.1\orca.exe")
# ========================================================

INP_DIR = ROOT / "INP"
OUT_DIR = ROOT / "OUT"
MID_DIR = ROOT / "MID"


def make_run_stamp() -> str:
    """Readable folder name like 'Apr. 02 2026 10.37am'."""
    now = datetime.now()
    # %b gives 'Apr', add period; %I.%M gives '10.37'; %p gives 'AM'
    raw = now.strftime("%b. %d %Y %I.%M%p")
    # Lowercase the AM/PM → 'am'/'pm'
    return raw[:-2] + raw[-2:].lower()


# ══════════════════════════════════════════════════════════════════
#  ORCA JOB RUNNER
# ══════════════════════════════════════════════════════════════════

def fmt_hhmmss(seconds: float) -> str:
    s = int(seconds)
    h, rem = divmod(s, 3600)
    m, sec = divmod(rem, 60)
    return f"{h:02d}:{m:02d}:{sec:02d}"


def usage(exit_code: int = 2):
    prog = Path(sys.argv[0]).name
    print("Usage:")
    print(f"  {prog} jobname [jobname2 ...]")
    print(f"  {prog} jobname.inp other.inp")
    print(rf'  {prog} "C:\full\path\to\file1.inp" "C:\full\path\to\file2.inp"')
    print(f"  {prog} -all              Run every .inp file in the INP folder")
    print()
    print(f'Inputs folder:  "{INP_DIR}"')
    print(f'Outputs folder: "{OUT_DIR}"')
    print(f'Mid folder:     "{MID_DIR}"')
    sys.exit(exit_code)


def resolve_input(arg: str):
    a = arg.strip().strip('"')
    ap = Path(a)
    if ap.exists():
        return ap, ap.stem
    if (INP_DIR / a).exists():
        return INP_DIR / a, Path(a).stem
    if (INP_DIR / f"{a}.inp").exists():
        return (INP_DIR / f"{a}.inp"), a
    if (INP_DIR / f"{a}.INP").exists():
        return (INP_DIR / f"{a}.INP"), a
    return None, None


def collect_all_inp() -> list:
    files = sorted(INP_DIR.glob("*.inp"), key=lambda p: p.name.lower())
    upper = sorted(INP_DIR.glob("*.INP"), key=lambda p: p.name.lower())
    seen = {p.name.lower() for p in files}
    for p in upper:
        if p.name.lower() not in seen:
            files.append(p)
            seen.add(p.name.lower())
    files.sort(key=lambda p: p.name.lower())
    return [(p.stem, p) for p in files]


def run_one_job(
    base: str,
    inp_src: Path,
    mid_run_dir: Path,
    out_run_dir: Path,
    env: dict,
    total_start: float,
    idx: int,
    n_jobs: int,
) -> int:
    """Run a single ORCA job. Returns 0 on success, >0 on error."""
    job_dir     = mid_run_dir / base
    job_out_dir = out_run_dir / base
    out_file    = job_out_dir / f"{base}.out"
    err_file    = job_out_dir / f"{base}.out.err"

    job_dir.mkdir(parents=True, exist_ok=True)
    job_out_dir.mkdir(parents=True, exist_ok=True)

    dst_inp = job_dir / f"{base}.inp"
    try:
        shutil.copy2(inp_src, dst_inp)
    except Exception:
        print(f'ERROR: Failed to copy input to "{job_dir}"')
        return 1

    tmpdir = job_dir / "_tmp"
    tmpdir.mkdir(parents=True, exist_ok=True)

    env_local = dict(env)
    env_local["TMPDIR"] = str(tmpdir)
    env_local["TEMP"]   = str(tmpdir)
    env_local["TMP"]    = str(tmpdir)

    print()
    print(f"Job:     {base}")
    print(f"Input:   {inp_src}")
    print(f"Workdir: {job_dir}")
    print(f"Out:     {out_file}")
    print(f"ORCA:    {ORCA_EXE}")
    print("-" * 60)

    job_start = time.time()

    with out_file.open("wb") as fout, err_file.open("wb") as ferr:
        try:
            p = subprocess.Popen(
                [str(ORCA_EXE), dst_inp.name],
                cwd=str(job_dir),
                stdout=fout,
                stderr=ferr,
                env=env_local,
                shell=False,
            )
        except Exception as e:
            print(f"ERROR: Failed to start ORCA: {e}")
            return 1

        while p.poll() is None:
            job_elapsed  = fmt_hhmmss(time.time() - job_start)
            total_elapsed = fmt_hhmmss(time.time() - total_start)
            line = f"[{idx}/{n_jobs}] {base} | Job {job_elapsed} | Total {total_elapsed}"
            print("\r" + line.ljust(100), end="", flush=True)
            time.sleep(1)

        rc = p.returncode
        job_elapsed  = fmt_hhmmss(time.time() - job_start)
        total_elapsed = fmt_hhmmss(time.time() - total_start)
        line = f"[{idx}/{n_jobs}] {base} | Job {job_elapsed} | Total {total_elapsed}  finished"
        print("\r" + line.ljust(100))

    # Append stderr to .out, then delete .err
    try:
        if err_file.exists():
            with err_file.open("rb") as ferr, out_file.open("ab") as fout:
                fout.write(b"\n")
                fout.write(ferr.read())
            try:
                err_file.unlink()
            except OSError:
                pass
    except Exception:
        pass

    # Copy .xyz files from MID to the output subfolder
    for xyz in job_dir.glob("*.xyz"):
        try:
            shutil.copy2(xyz, job_out_dir / xyz.name)
        except Exception:
            pass

    if rc != 0:
        print(f'ORCA exited with code {rc}. See: "{out_file}"')

    return rc


# ══════════════════════════════════════════════════════════════════
#  OUTPUT PARSER + EXCEL REPORT
# ══════════════════════════════════════════════════════════════════

HEADER_FILL = PatternFill("solid", fgColor="2F5496") if HAS_OPENPYXL else None
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11) if HAS_OPENPYXL else None
ERROR_FILL  = PatternFill("solid", fgColor="FFC7CE") if HAS_OPENPYXL else None
ERROR_FONT  = Font(name="Arial", color="9C0006") if HAS_OPENPYXL else None
NORMAL_FONT = Font(name="Arial", size=10) if HAS_OPENPYXL else None
BOLD_FONT   = Font(name="Arial", size=10, bold=True) if HAS_OPENPYXL else None
THIN_BORDER = Border(bottom=Side(style="thin", color="D9D9D9")) if HAS_OPENPYXL else None

NUM_FMT_EH   = '0.00000000" Eh"'
NUM_FMT_KCAL = '0.00" kcal/mol"'
NUM_FMT_SEC  = '#,##0.000" sec"'
NUM_FMT_PCT  = '0.0"%"'


def parse_out_file(path: Path) -> dict:
    text = path.read_text(errors="replace")
    result = {
        "file": path.name,
        "normal_term": False,
        "errors": [],
        "enthalpy_eh": None,
        "entropy_eh": None,
        "entropy_kcal": None,
        "gibbs_eh": None,
        "timings": [],
        "total_run_time": None,
    }

    if "****ORCA TERMINATED NORMALLY****" in text:
        result["normal_term"] = True

    for line in text.splitlines():
        stripped = line.strip()
        if stripped.startswith("ORCA ABORT") or stripped.startswith("ERROR"):
            result["errors"].append(stripped)
        elif "ABORTING THE RUN" in stripped:
            result["errors"].append(stripped)

    if not result["normal_term"] and not result["errors"]:
        result["errors"].append("Job did not terminate normally (no specific error captured)")

    m = re.search(r"Total enthalpy\s+\.\.\.\s+([-\d.]+)\s*Eh", text)
    if m:
        result["enthalpy_eh"] = float(m.group(1))

    m = re.search(
        r"Total entropy correction\s+\.\.\.\s+([-\d.]+)\s*Eh\s+([-\d.]+)\s*kcal/mol", text
    )
    if m:
        result["entropy_eh"]   = float(m.group(1))
        result["entropy_kcal"] = float(m.group(2))

    m = re.search(r"Final Gibbs free energy\s+\.\.\.\s+([-\d.]+)\s*Eh", text)
    if m:
        result["gibbs_eh"] = float(m.group(1))

    timing_pat = re.compile(
        r"^([\w\s]+?)\s+\.{3}\s+([\d.]+)\s+sec\s+\(=\s+[\d.]+\s+min\)\s+([\d.]+)\s*%",
        re.MULTILINE,
    )
    for tm in timing_pat.finditer(text):
        result["timings"].append({
            "module": tm.group(1).strip(),
            "seconds": float(tm.group(2)),
            "pct": float(tm.group(3)),
        })

    m = re.search(r"Sum of individual times\s+\.\.\.\s+([\d.]+)\s+sec", text)
    if m:
        result["timings"].insert(0, {
            "module": "Sum of individual times",
            "seconds": float(m.group(1)),
            "pct": None,
        })

    m = re.search(r"TOTAL RUN TIME:\s*(.+)", text)
    if m:
        result["total_run_time"] = m.group(1).strip()

    return result


def style_header_row(ws, row, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")


def auto_width(ws, min_width=10, max_width=45):
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        lengths = [len(str(c.value)) for c in col_cells if c.value is not None]
        best = max(lengths) if lengths else min_width
        ws.column_dimensions[col_letter].width = min(max(best + 3, min_width), max_width)


def build_report(out_files: list[Path], xlsx_path: Path):
    if not HAS_OPENPYXL:
        print("\n[WARNING] openpyxl not installed — skipping Excel report.")
        print("  Install it with:  pip install openpyxl")
        return

    if not out_files:
        print("\nNo .out files to report on.")
        return

    records = []
    for f in out_files:
        if f.exists():
            records.append(parse_out_file(f))

    if not records:
        return

    wb = Workbook()

    # ── Summary sheet ─────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    headers = [
        "File", "Normal Termination?",
        "Total Enthalpy (Eh)", "Entropy Correction (Eh)",
        "Entropy Correction (kcal/mol)", "Final Gibbs Free Energy (Eh)",
        "Total Run Time", "Errors",
    ]
    ws.append(headers)
    style_header_row(ws, 1, len(headers))
    ws.freeze_panes = "A2"

    for r in records:
        err_str = "; ".join(r["errors"]) if r["errors"] else ""
        ws.append([
            r["file"],
            "YES" if r["normal_term"] else "NO",
            r["enthalpy_eh"],
            r["entropy_eh"],
            r["entropy_kcal"],
            r["gibbs_eh"],
            r["total_run_time"] or "",
            err_str,
        ])
        row_idx = ws.max_row

        for col, fmt in [(3, NUM_FMT_EH), (4, NUM_FMT_EH),
                         (5, NUM_FMT_KCAL), (6, NUM_FMT_EH)]:
            c = ws.cell(row=row_idx, column=col)
            if c.value is not None:
                c.number_format = fmt

        if not r["normal_term"]:
            for col in range(1, len(headers) + 1):
                c = ws.cell(row=row_idx, column=col)
                c.fill = ERROR_FILL
                c.font = ERROR_FONT

        for col in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=col).border = THIN_BORDER

    auto_width(ws)

    # ── Timings sheet ─────────────────────────────────────────
    ws2 = wb.create_sheet("Timings")
    t_headers = ["File", "Module", "Time (sec)", "Percent (%)"]
    ws2.append(t_headers)
    style_header_row(ws2, 1, len(t_headers))
    ws2.freeze_panes = "A2"

    for r in records:
        for t in r["timings"]:
            ws2.append([r["file"], t["module"], t["seconds"], t["pct"]])
            row_idx = ws2.max_row
            ws2.cell(row=row_idx, column=3).number_format = NUM_FMT_SEC
            if t["pct"] is not None:
                ws2.cell(row=row_idx, column=4).number_format = NUM_FMT_PCT
            for col in range(1, len(t_headers) + 1):
                ws2.cell(row=row_idx, column=col).border = THIN_BORDER
                ws2.cell(row=row_idx, column=col).font = NORMAL_FONT

        if r["total_run_time"]:
            ws2.append([r["file"], "TOTAL RUN TIME", r["total_run_time"], ""])
            row_idx = ws2.max_row
            for col in range(1, len(t_headers) + 1):
                c = ws2.cell(row=row_idx, column=col)
                c.font = BOLD_FONT
                c.border = THIN_BORDER

        ws2.append([])

    auto_width(ws2)

    # ── Errors sheet (only if needed) ─────────────────────────
    if any(r["errors"] for r in records):
        ws3 = wb.create_sheet("Errors")
        e_headers = ["File", "Error Message"]
        ws3.append(e_headers)
        style_header_row(ws3, 1, len(e_headers))
        ws3.freeze_panes = "A2"

        for r in records:
            for e in r["errors"]:
                ws3.append([r["file"], e])
                row_idx = ws3.max_row
                for col in range(1, len(e_headers) + 1):
                    c = ws3.cell(row=row_idx, column=col)
                    c.fill = ERROR_FILL
                    c.font = ERROR_FONT
                    c.border = THIN_BORDER
        auto_width(ws3, max_width=80)

    wb.save(xlsx_path)
    print(f"\nExcel report saved: {xlsx_path}")
    print(f"  Summary tab:  Gibbs free energies for all jobs")
    print(f"  Timings tab:  module-level timing breakdown")
    if any(r["errors"] for r in records):
        print(f"  Errors tab:   error messages from failed jobs")


# ══════════════════════════════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════════════════════════════

def main():
    if len(sys.argv) < 2:
        usage(2)

    if not ORCA_EXE.exists():
        print(f'ERROR: orca.exe not found at "{ORCA_EXE}"')
        return 1
    if not INP_DIR.exists():
        print(f'ERROR: INP folder not found: "{INP_DIR}"')
        return 1

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    MID_DIR.mkdir(parents=True, exist_ok=True)

    # Handle -all flag
    if len(sys.argv) == 2 and sys.argv[1].lower() in ("-all", "--all"):
        jobs = collect_all_inp()
        if not jobs:
            print(f'No .inp files found in "{INP_DIR}"')
            return 1
        print(f"Found {len(jobs)} .inp file(s) in INP folder:")
        for base, inp_src in jobs:
            print(f"  {inp_src.name}")
    else:
        jobs = []
        missing = []
        for raw in sys.argv[1:]:
            inp_src, base = resolve_input(raw)
            if inp_src is None:
                missing.append(raw)
            else:
                jobs.append((base, inp_src))

        if missing:
            print("ERROR: One or more inputs were not found:")
            for m in missing:
                print(f"  {m}")
            return 1

    # collision check
    seen = {}
    dups = []
    for base, inp_src in jobs:
        k = base.lower()
        if k in seen:
            dups.append((base, seen[k], inp_src))
        else:
            seen[k] = inp_src

    if dups:
        print("ERROR: Overlapping job names detected (would collide in MID/OUT):")
        for base, first_src, dup_src in dups:
            print(f'  base="{base}" from:')
            print(f"    - {first_src}")
            print(f"    - {dup_src}")
        return 1

    # Create matching timestamped folders in OUT and MID
    stamp = make_run_stamp()
    out_run_dir = OUT_DIR / stamp
    mid_run_dir = MID_DIR / stamp
    out_run_dir.mkdir(parents=True, exist_ok=True)
    mid_run_dir.mkdir(parents=True, exist_ok=True)

    env = os.environ.copy()
    total_start = time.time()

    print()
    print(f"Run:        {stamp}")
    print(f"Queue size: {len(jobs)} job(s)")
    print("-" * 60)

    results = []
    for i, (base, inp_src) in enumerate(jobs, start=1):
        rc = run_one_job(base, inp_src, mid_run_dir, out_run_dir,
                         env, total_start, i, len(jobs))
        results.append((base, rc))
        if rc != 0:
            total_elapsed = fmt_hhmmss(time.time() - total_start)
            print(f"\nQueue stopped due to failure in job '{base}'. Total elapsed: {total_elapsed}")
            break

    total_elapsed = fmt_hhmmss(time.time() - total_start)
    print()
    print("-" * 60)
    print(f"All jobs finished. Total elapsed: {total_elapsed}")
    print("Summary:")
    for base, rc in results:
        print(f"  {base}: rc={rc}")

    # ── Generate Excel report ─────────────────────────────────
    out_files = [out_run_dir / base / f"{base}.out" for base, _ in results]
    xlsx_path = out_run_dir / "orca_report.xlsx"
    build_report(out_files, xlsx_path)

    return 0


if __name__ == "__main__":
    raise SystemExit(main())