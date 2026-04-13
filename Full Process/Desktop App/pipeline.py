"""
pipeline.py - Core pipeline: SMILES -> 3D -> ORCA input -> run -> parse -> report

Handles the full computational chemistry workflow:
  1. Convert SMILES to 3D geometries via RDKit force-field optimization
  2. Generate ORCA input files from settings + geometry
  3. Run ORCA calculations
  4. Parse .out files for thermodynamic data
  5. Build Excel report with embedded molecule structure images
"""

import io
import os
import re
import sys
import time
import shutil
import subprocess
import threading
from pathlib import Path
from datetime import datetime

import numpy as np

from rdkit import Chem, RDLogger
from rdkit.Chem import AllChem, Descriptors, Draw, rdMolAlign
from PIL import Image as PILImage

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XlImage

ROOT = Path(__file__).resolve().parent

# ===================== USER SETTING =====================
ORCA_EXE = Path(r"C:\ORCA_6.1.1\orca.exe")
# ========================================================

INP_DIR = ROOT / "INP"
OUT_DIR = ROOT / "OUT"
MID_DIR = ROOT / "MID"


def make_run_stamp() -> str:
    """Readable folder name like 'Apr. 02 2026 10.37am'."""
    now = datetime.now()
    raw = now.strftime("%b. %d %Y %I.%M%p")
    return raw[:-2] + raw[-2:].lower()


# -------------------------------------------------------------------
#  SMILES -> 3D GEOMETRY + MOLECULE IMAGE
# -------------------------------------------------------------------

def validate_smiles(smiles: str) -> bool:
    """Return True if RDKit can parse the SMILES string."""
    return Chem.MolFromSmiles(smiles) is not None


def smiles_to_xyz(smiles: str, num_confs: int = 50):
    """Convert a SMILES string to optimized 3D coordinates and a 2D image.

    Uses RDKit ETKDG multi-conformer generation: embeds *num_confs*
    conformers, MMFF-optimises each, and picks the lowest-energy one.

    Returns:
        (xyz_block, png_bytes, num_atoms)
    Raises:
        ValueError on invalid SMILES or embedding failure.
    """
    mol = Chem.MolFromSmiles(smiles)
    if mol is None:
        raise ValueError(f"Invalid SMILES: {smiles}")

    mol_3d = Chem.AddHs(mol)

    # Multi-conformer embedding with ETKDGv3
    params = AllChem.ETKDGv3()
    params.randomSeed = 42
    params.numThreads = 0  # use all available cores
    conf_ids = AllChem.EmbedMultipleConfs(mol_3d, numConfs=num_confs, params=params)

    if len(conf_ids) == 0:
        # Fallback: single conformer with basic embedding
        if AllChem.EmbedMolecule(mol_3d, randomSeed=42) == -1:
            raise ValueError(f"3D embedding failed for: {smiles}")
        conf_ids = [0]

    # MMFF-optimise each conformer and pick the lowest energy
    results = AllChem.MMFFOptimizeMoleculeConfs(mol_3d, maxIters=2000)

    best_conf = 0
    best_energy = float("inf")
    for cid, (converged, energy) in zip(conf_ids, results):
        if energy < best_energy:
            best_energy = energy
            best_conf = cid

    # Build XYZ coordinate block from the best conformer
    conf = mol_3d.GetConformer(best_conf)
    lines = []
    for i, atom in enumerate(mol_3d.GetAtoms()):
        pos = conf.GetAtomPosition(i)
        lines.append(
            f"  {atom.GetSymbol():2s}  {pos.x:14.8f}  {pos.y:14.8f}  {pos.z:14.8f}"
        )

    xyz_block = "\n".join(lines)
    num_atoms = mol_3d.GetNumAtoms()

    # Generate 2D structure image (PNG)
    mol_2d = Chem.MolFromSmiles(smiles)
    AllChem.Compute2DCoords(mol_2d)
    img = Draw.MolToImage(mol_2d, size=(400, 300))
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    png_bytes = buf.getvalue()

    return xyz_block, png_bytes, num_atoms


def _consolidate_fragments(mol_3d):
    """Move fragments close together so inter-fragment FF interactions work.

    After EmbedMolecule, disconnected fragments (salt pairs, ion complexes)
    can be placed very far apart.  This translates each fragment so its
    centroid is within ~3 Å of the largest fragment's centroid.
    """
    frags = Chem.GetMolFrags(mol_3d, asMols=False)
    if len(frags) <= 1:
        return  # single fragment, nothing to do

    conf = mol_3d.GetConformer()

    # Find centroid of each fragment
    centroids = []
    for frag_idxs in frags:
        pts = np.array([list(conf.GetAtomPosition(i)) for i in frag_idxs])
        centroids.append(pts.mean(axis=0))

    # Largest fragment stays put; move others near it
    largest_idx = max(range(len(frags)), key=lambda i: len(frags[i]))
    target = centroids[largest_idx]

    for fi, frag_idxs in enumerate(frags):
        if fi == largest_idx:
            continue
        # Offset to place this fragment's centroid ~3 Å from the target
        direction = centroids[fi] - target
        dist = np.linalg.norm(direction)
        if dist < 0.01:
            direction = np.array([1.0, 0.0, 0.0])
            dist = 1.0
        desired_dist = 3.0
        shift = target + direction / dist * desired_dist - centroids[fi]
        for i in frag_idxs:
            p = conf.GetAtomPosition(i)
            conf.SetAtomPosition(i, (p.x + shift[0], p.y + shift[1], p.z + shift[2]))


def smiles_to_mol3d(smiles: str):
    """Convert SMILES to a 3D mol object *without* force-field optimisation.

    Use this for the 3D viewer so the user sees the raw embedded geometry
    before deciding to run FF optimisation manually.

    Returns:
        (mol_3d, png_bytes, num_atoms)
    Raises:
        ValueError on invalid SMILES or embedding failure.
    """
    mol = Chem.MolFromSmiles(smiles)
    if mol is None:
        raise ValueError(f"Invalid SMILES: {smiles}")

    mol_3d = Chem.AddHs(mol)

    params = AllChem.ETKDGv3()
    params.randomSeed = 42
    if AllChem.EmbedMolecule(mol_3d, params) == -1:
        if AllChem.EmbedMolecule(mol_3d, randomSeed=42) == -1:
            raise ValueError(f"3D embedding failed for: {smiles}")

    # Bring disconnected fragments close together for proper FF interaction
    _consolidate_fragments(mol_3d)

    # 2D structure image
    mol_2d = Chem.MolFromSmiles(smiles)
    AllChem.Compute2DCoords(mol_2d)
    img = Draw.MolToImage(mol_2d, size=(400, 300))
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    png_bytes = buf.getvalue()

    return mol_3d, png_bytes, mol_3d.GetNumAtoms()


def ff_optimize_mol(mol_3d, num_confs: int = 50):
    """Run ETKDG multi-conformer search + MMFF optimisation on *mol_3d*.

    Generates *num_confs* conformers, MMFF-optimises each, and keeps the
    lowest-energy conformer.  Returns the same mol object.
    """
    # Generate additional conformers
    params = AllChem.ETKDGv3()
    params.randomSeed = 42
    params.numThreads = 0
    conf_ids = AllChem.EmbedMultipleConfs(mol_3d, numConfs=num_confs, params=params)

    if len(conf_ids) == 0:
        # Molecule already has a conformer from initial embed; just MMFF it
        try:
            AllChem.MMFFOptimizeMolecule(mol_3d, maxIters=2000)
        except Exception:
            try:
                AllChem.UFFOptimizeMolecule(mol_3d, maxIters=2000)
            except Exception:
                pass
        return mol_3d

    # MMFF-optimise all conformers and pick lowest energy
    results = AllChem.MMFFOptimizeMoleculeConfs(mol_3d, maxIters=2000)

    best_conf = conf_ids[0]
    best_energy = float("inf")
    for cid, (converged, energy) in zip(conf_ids, results):
        if energy < best_energy:
            best_energy = energy
            best_conf = cid

    # Keep only the best conformer
    if best_conf != 0:
        # Copy best conformer coords to conformer 0
        best_pos = mol_3d.GetConformer(best_conf)
        conf0 = mol_3d.GetConformer(0)
        for i in range(mol_3d.GetNumAtoms()):
            conf0.SetAtomPosition(i, best_pos.GetAtomPosition(i))

    # Remove all conformers except 0
    confs_to_remove = [c.GetId() for c in mol_3d.GetConformers() if c.GetId() != 0]
    for cid in confs_to_remove:
        mol_3d.RemoveConformer(cid)

    return mol_3d


def mol_to_xyz_block(mol_3d) -> str:
    """Extract an XYZ coordinate block from an RDKit mol with a 3D conformer."""
    conf = mol_3d.GetConformer()
    lines = []
    for i, atom in enumerate(mol_3d.GetAtoms()):
        pos = conf.GetAtomPosition(i)
        lines.append(
            f"  {atom.GetSymbol():2s}  {pos.x:14.8f}  {pos.y:14.8f}  {pos.z:14.8f}"
        )
    return "\n".join(lines)


def global_conformer_search(
    mol_3d,
    n_conformers: int | None = None,
    rmsd_threshold: float = 0.5,
    random_seed: int = -1,
) -> tuple[object, float, int, int, str]:
    """Stochastic global conformer search with RMSD deduplication.

    Generates many random starting geometries via ETKDG, MMFF-optimises
    all of them in parallel (one conformer per CPU core), deduplicates
    by RMSD using a fast C++ bulk matrix, and returns the lowest-energy
    unique conformer.

    Parameters
    ----------
    mol_3d : rdkit.Chem.Mol
        Molecule with explicit hydrogens (Chem.AddHs already applied).
    n_conformers : int or None
        Number of conformers to embed.  If None, auto-scales from the
        number of rotatable bonds: max(50, min(500, 50 * n_rot)).
    rmsd_threshold : float
        Minimum RMSD (Angstroms) between kept conformers.
    random_seed : int
        Seed for ETKDG.  Use -1 for non-deterministic.

    Returns
    -------
    (mol_result, best_energy_kcal, n_unique, n_cpus, ff_name)
        *mol_result* has a single conformer (the global minimum).
        *best_energy_kcal* is the force-field energy in kcal/mol.
        *n_unique* is how many unique conformers passed the RMSD filter.
        *n_cpus* is the number of CPU threads used.
        *ff_name* is "MMFF" or "UFF" depending on which force field was used.

    Raises
    ------
    ValueError
        If embedding or force-field optimisation fails entirely.
    """
    n_cpus = os.cpu_count() or 4

    if n_conformers is None:
        n_rot = Descriptors.NumRotatableBonds(mol_3d)
        n_conformers = max(50, min(500, 50 * n_rot))

    # --- Phase 1: parallel embedding (each conformer on its own core) ---
    params = AllChem.ETKDGv3()
    params.randomSeed = random_seed
    params.numThreads = n_cpus
    params.useRandomCoords = True

    conf_ids = list(AllChem.EmbedMultipleConfs(
        mol_3d, numConfs=n_conformers, params=params,
    ))

    if len(conf_ids) == 0:
        if AllChem.EmbedMolecule(mol_3d, randomSeed=42) == -1:
            raise ValueError("3D embedding failed — cannot generate any conformers.")
        conf_ids = [0]

    # --- Phase 2: parallel FF optimisation (one conformer per core) ---
    # Suppress RDKit valence warnings (e.g. charged O with 3 bonds)
    # that are non-fatal but noisy for unusual molecules.
    RDLogger.DisableLog("rdApp.warning")
    try:
        results = None
        ff_name = "MMFF"
        try:
            results = AllChem.MMFFOptimizeMoleculeConfs(
                mol_3d, maxIters=2000, numThreads=n_cpus,
            )
        except Exception:
            results = None

        # If MMFF failed entirely (or every conformer returned -1), use UFF
        if results is None or all(r[0] == -1 for r in results):
            ff_name = "UFF"
            try:
                results = AllChem.UFFOptimizeMoleculeConfs(
                    mol_3d, maxIters=2000, numThreads=n_cpus,
                )
            except Exception:
                raise ValueError(
                    "Both MMFF and UFF force fields failed for this molecule.\n"
                    "This usually means the molecule has unusual valence states."
                )
    finally:
        RDLogger.EnableLog("rdApp.warning")

    # Build (conf_id, energy) pairs, skipping failures
    energies = []
    for cid, (converged, energy) in zip(conf_ids, results):
        if converged != -1:  # -1 means FF setup failed for that conformer
            energies.append((cid, energy))

    if not energies:
        raise ValueError("Force-field optimisation failed for all conformers.")

    # Sort by energy (ascending)
    energies.sort(key=lambda x: x[1])

    # --- Phase 3: fast bulk RMSD deduplication (C++ matrix) ---
    n_total = mol_3d.GetNumConformers()
    if n_total > 1:
        # Build conformer-id -> sequential-index mapping
        conf_list = [c.GetId() for c in mol_3d.GetConformers()]
        id_to_idx = {cid: i for i, cid in enumerate(conf_list)}

        # Compute all pairwise RMSDs in one C++ call (much faster than
        # calling GetBestRMS in a Python loop).
        rms_matrix = AllChem.GetConformerRMSMatrix(mol_3d, prealigned=False)

        def _rms_lookup(ci: int, cj: int) -> float:
            """Look up RMSD from the condensed distance matrix."""
            if ci == cj:
                return 0.0
            if ci > cj:
                ci, cj = cj, ci
            idx = ci * n_total - ci * (ci + 1) // 2 + cj - ci - 1
            return rms_matrix[idx]

        # Greedy deduplication: walk energy-sorted list, keep a conformer
        # only if its RMSD to every already-kept conformer >= threshold.
        kept: list[tuple[int, float]] = []
        for cid, energy in energies:
            ci = id_to_idx[cid]
            is_unique = True
            for kept_cid, _ in kept:
                ki = id_to_idx[kept_cid]
                if _rms_lookup(ci, ki) < rmsd_threshold:
                    is_unique = False
                    break
            if is_unique:
                kept.append((cid, energy))
    else:
        kept = energies[:1]

    if not kept:
        kept = [energies[0]]

    best_cid, best_energy = kept[0]

    # Copy best conformer coords to conformer 0, remove all others
    if best_cid != 0:
        best_pos = mol_3d.GetConformer(best_cid)
        conf0 = mol_3d.GetConformer(0)
        for i in range(mol_3d.GetNumAtoms()):
            conf0.SetAtomPosition(i, best_pos.GetAtomPosition(i))

    confs_to_remove = [c.GetId() for c in mol_3d.GetConformers() if c.GetId() != 0]
    for cid in confs_to_remove:
        mol_3d.RemoveConformer(cid)

    return mol_3d, best_energy, len(kept), n_cpus, ff_name


# -------------------------------------------------------------------
#  GOAT CONFORMER SEARCH
# -------------------------------------------------------------------

def generate_goat_inp(xyz_block: str, settings: dict) -> str:
    """Build an ORCA input file for a GOAT global conformer search.

    Uses GFN2-xTB (fast semi-empirical method) with GOAT global optimizer.
    Charge, multiplicity, and core count are taken from the same settings
    as the DFT step.
    """
    charge = settings.get("charge", 0)
    multiplicity = settings.get("multiplicity", 1)
    cpus = settings.get("cpus", 4)
    ram_mb = settings.get("ram_mb", 4000)

    parts = [
        "! XTB2 GOAT",
        "",
        f"%maxcore {ram_mb}",
        "",
        "%pal",
        f"  nprocs {cpus}",
        "end",
        "",
        "%goat",
        f"  NWorkers {cpus}",
        "end",
        "",
        f"* xyz {charge} {multiplicity}",
        xyz_block,
        "*",
        "",
    ]
    return "\n".join(parts)


def run_goat_job(
    base: str,
    xyz_block: str,
    settings: dict,
    job_dir: Path,
    env: dict,
    status_dict: dict | None = None,
    pipeline_start: float | None = None,
    orca_exe: Path | None = None,
    abort_event: threading.Event | None = None,
) -> tuple[str | None, str | None]:
    """Run a GOAT conformer search and return the global-minimum geometry.

    Creates a ``goat_preopt/`` subdirectory inside *job_dir*, writes and
    executes the GOAT input there, then extracts coordinates from the
    ``.globalminimum.xyz`` file ORCA produces.

    Returns
    -------
    (xyz_block, warning)
        *xyz_block* is the global-minimum geometry (or None on failure).
        *warning* is a message if GOAT failed (caller should fall back to
        the original RDKit geometry).
    """
    orca_exe = orca_exe or ORCA_EXE
    goat_dir = job_dir / "goat_preopt"
    goat_dir.mkdir(parents=True, exist_ok=True)

    goat_name = f"{base}_goat"
    inp_content = generate_goat_inp(xyz_block, settings)
    inp_path = goat_dir / f"{goat_name}.inp"
    inp_path.write_text(inp_content, encoding="utf-8")

    out_file = goat_dir / f"{goat_name}.out"
    err_file = goat_dir / f"{goat_name}.out.err"

    tmpdir = goat_dir / "_tmp"
    tmpdir.mkdir(parents=True, exist_ok=True)

    env_local = dict(env)
    env_local["TMPDIR"] = str(tmpdir)
    env_local["TEMP"] = str(tmpdir)
    env_local["TMP"] = str(tmpdir)

    creationflags = 0
    if sys.platform == "win32":
        creationflags = subprocess.CREATE_NO_WINDOW

    with out_file.open("wb") as fout, err_file.open("wb") as ferr:
        try:
            p = subprocess.Popen(
                [str(orca_exe), inp_path.name],
                cwd=str(goat_dir),
                stdout=fout,
                stderr=ferr,
                env=env_local,
                shell=False,
                creationflags=creationflags,
            )
        except Exception as exc:
            return None, f"GOAT failed to start: {exc}"

        while p.poll() is None:
            if abort_event is not None and abort_event.is_set():
                p.kill()
                p.wait()
                return None, "GOAT aborted by user"
            if status_dict is not None and pipeline_start is not None:
                status_dict["elapsed"] = fmt_hhmmss(time.monotonic() - pipeline_start)
            time.sleep(2)

    # Append stderr into .out
    if err_file.exists():
        try:
            with err_file.open("rb") as ef, out_file.open("ab") as of:
                of.write(b"\n")
                of.write(ef.read())
            err_file.unlink()
        except OSError:
            pass

    # Check for normal termination
    try:
        out_text = out_file.read_text(errors="replace")
        normal = "****ORCA TERMINATED NORMALLY****" in out_text
    except Exception:
        normal = False

    if not normal:
        return None, "GOAT conformer search failed. Falling back to RDKit geometry."

    # Parse .globalminimum.xyz
    gmin_xyz = goat_dir / f"{goat_name}.globalminimum.xyz"
    if not gmin_xyz.exists():
        return None, "GOAT conformer search failed. Falling back to RDKit geometry."

    goat_xyz = parse_goat_xyz(gmin_xyz)
    if goat_xyz is None:
        return None, "GOAT conformer search failed. Falling back to RDKit geometry."

    # Copy the globalminimum.xyz up to the main job directory
    try:
        shutil.copy2(gmin_xyz, job_dir / f"{base}.globalminimum.xyz")
    except Exception:
        pass

    return goat_xyz, None


def parse_goat_xyz(path: Path) -> str | None:
    """Parse a standard .xyz file and return an XYZ coordinate block.

    Standard XYZ format:
        Line 1: atom count
        Line 2: comment
        Lines 3+: element  x  y  z

    Returns the coordinate block formatted for ORCA input, or None on error.
    """
    try:
        lines = path.read_text(encoding="utf-8").strip().splitlines()
        if len(lines) < 3:
            return None
        coord_lines = []
        for line in lines[2:]:
            parts = line.split()
            if len(parts) >= 4:
                elem = parts[0]
                x, y, z = float(parts[1]), float(parts[2]), float(parts[3])
                coord_lines.append(f"  {elem:2s}  {x:14.8f}  {y:14.8f}  {z:14.8f}")
        return "\n".join(coord_lines) if coord_lines else None
    except Exception:
        return None


# -------------------------------------------------------------------
#  ORCA INPUT FILE GENERATION
# -------------------------------------------------------------------

def generate_inp(name: str, xyz_block: str, settings: dict) -> str:
    """Build ORCA .inp file content from XYZ coordinates and user settings."""
    functional = settings.get("functional", "B3LYP")
    basis_set = settings.get("basis_set", "def2-SVP")
    calc_type = settings.get("calc_type", "OPT FREQ")
    charge = settings.get("charge", 0)
    multiplicity = settings.get("multiplicity", 1)
    ram_mb = settings.get("ram_mb", 4000)
    cpus = settings.get("cpus", 4)
    extra_keywords = settings.get("extra_keywords", "").strip()
    extra_blocks = settings.get("extra_blocks", "").strip()
    nprocs_group = settings.get("nprocs_group")

    kw_parts = [functional, basis_set]
    if calc_type:
        kw_parts.append(calc_type)
    keyword_line = "! " + " ".join(kw_parts)
    if extra_keywords:
        keyword_line += f" {extra_keywords}"

    parts = [
        keyword_line,
        "",
        f"%maxcore {ram_mb}",
        "%pal",
        f"  nprocs {cpus}",
    ]
    if nprocs_group is not None:
        parts.append(f"  nprocs_group {nprocs_group}")
    parts.append("end")

    if extra_blocks:
        parts += ["", extra_blocks]

    parts += [
        "",
        f"* xyz {charge} {multiplicity}",
        xyz_block,
        "*",
        "",
    ]

    return "\n".join(parts)


# -------------------------------------------------------------------
#  ORCA JOB RUNNER
# -------------------------------------------------------------------

def fmt_hhmmss(seconds: float) -> str:
    s = int(seconds)
    h, rem = divmod(s, 3600)
    m, sec = divmod(rem, 60)
    return f"{h:02d}:{m:02d}:{sec:02d}"


def run_orca_job(
    base: str,
    inp_src: Path,
    mid_dir: Path,
    out_dir: Path,
    env: dict,
    status_dict: dict | None = None,
    pipeline_start: float | None = None,
    orca_exe: Path | None = None,
    abort_event: threading.Event | None = None,
) -> int:
    """Run a single ORCA job.  Returns process exit code (0 = success).

    If *status_dict* and *pipeline_start* are provided the function
    periodically updates status_dict["elapsed"] so the web UI can
    show a live timer.  An optional *orca_exe* overrides the module-level
    ORCA_EXE path.  If *abort_event* is set the process is killed and
    -9 is returned.
    """
    orca_exe = orca_exe or ORCA_EXE
    job_dir = mid_dir / base
    job_out_dir = out_dir / base
    out_file = job_out_dir / f"{base}.out"
    err_file = job_out_dir / f"{base}.out.err"

    job_dir.mkdir(parents=True, exist_ok=True)
    job_out_dir.mkdir(parents=True, exist_ok=True)

    dst_inp = job_dir / f"{base}.inp"
    shutil.copy2(inp_src, dst_inp)

    tmpdir = job_dir / "_tmp"
    tmpdir.mkdir(parents=True, exist_ok=True)

    env_local = dict(env)
    env_local["TMPDIR"] = str(tmpdir)
    env_local["TEMP"] = str(tmpdir)
    env_local["TMP"] = str(tmpdir)

    # On Windows, suppress the console window for child processes
    creationflags = 0
    if sys.platform == "win32":
        creationflags = subprocess.CREATE_NO_WINDOW

    with out_file.open("wb") as fout, err_file.open("wb") as ferr:
        try:
            p = subprocess.Popen(
                [str(orca_exe), dst_inp.name],
                cwd=str(job_dir),
                stdout=fout,
                stderr=ferr,
                env=env_local,
                shell=False,
                creationflags=creationflags,
            )
        except Exception:
            return 1

        # Poll so we can update the live timer and honour abort
        while p.poll() is None:
            if abort_event is not None and abort_event.is_set():
                p.kill()
                p.wait()
                return -9
            if status_dict is not None and pipeline_start is not None:
                status_dict["elapsed"] = fmt_hhmmss(time.monotonic() - pipeline_start)
            time.sleep(2)

        rc = p.returncode

    # Append stderr into .out, then delete .err
    if err_file.exists():
        try:
            with err_file.open("rb") as ef, out_file.open("ab") as of:
                of.write(b"\n")
                of.write(ef.read())
            err_file.unlink()
        except OSError:
            pass

    # Copy .xyz artefacts to the output subfolder
    for xyz in job_dir.glob("*.xyz"):
        try:
            shutil.copy2(xyz, job_out_dir / xyz.name)
        except Exception:
            pass

    return rc


# -------------------------------------------------------------------
#  ORCA OUTPUT PARSER
# -------------------------------------------------------------------

def parse_out_file(path: Path) -> dict:
    """Extract thermodynamic data, timings, and errors from an ORCA .out file."""
    text = path.read_text(errors="replace")

    result = {
        "file": path.name,
        "normal_term": False,
        "errors": [],
        "electronic_energy_eh": None,
        "enthalpy_eh": None,
        "entropy_eh": None,
        "entropy_kcal": None,
        "gibbs_eh": None,
        "timings": [],
        "total_run_time": None,
    }

    if "****ORCA TERMINATED NORMALLY****" in text:
        result["normal_term"] = True

    for line in text.splitlines():
        s = line.strip()
        if s.startswith("ORCA ABORT") or s.startswith("ERROR"):
            result["errors"].append(s)
        elif "ABORTING THE RUN" in s:
            result["errors"].append(s)

    if not result["normal_term"] and not result["errors"]:
        result["errors"].append("Job did not terminate normally (no specific error captured)")

    m = re.search(r"FINAL SINGLE POINT ENERGY\s+([-\d.]+)", text)
    if m:
        result["electronic_energy_eh"] = float(m.group(1))

    m = re.search(r"Total enthalpy\s+\.\.\.\s+([-\d.]+)\s*Eh", text)
    if m:
        result["enthalpy_eh"] = float(m.group(1))

    m = re.search(
        r"Total entropy correction\s+\.\.\.\s+([-\d.]+)\s*Eh\s+([-\d.]+)\s*kcal/mol",
        text,
    )
    if m:
        result["entropy_eh"] = float(m.group(1))
        result["entropy_kcal"] = float(m.group(2))

    m = re.search(r"Final Gibbs free energy\s+\.\.\.\s+([-\d.]+)\s*Eh", text)
    if m:
        result["gibbs_eh"] = float(m.group(1))

    m = re.search(r"TOTAL RUN TIME:\s*(.+)", text)
    if m:
        result["total_run_time"] = m.group(1).strip()

    timing_pat = re.compile(
        r"^([\w\s]+?)\s+\.{3}\s+([\d.]+)\s+sec\s+\(=\s+[\d.]+\s+min\)\s+([\d.]+)\s*%",
        re.MULTILINE,
    )
    for tm in timing_pat.finditer(text):
        result["timings"].append(
            {
                "module": tm.group(1).strip(),
                "seconds": float(tm.group(2)),
                "pct": float(tm.group(3)),
            }
        )

    return result


# -------------------------------------------------------------------
#  EXCEL REPORT WITH MOLECULE IMAGES
# -------------------------------------------------------------------

HEADER_FILL = PatternFill("solid", fgColor="1e3a5f")
HEADER_FONT = Font(name="Segoe UI", bold=True, color="FFFFFF", size=11)
NORMAL_FONT = Font(name="Segoe UI", size=10)
BOLD_FONT = Font(name="Segoe UI", size=10, bold=True)
ERROR_FILL = PatternFill("solid", fgColor="FFC7CE")
ERROR_FONT = Font(name="Segoe UI", color="9C0006", size=10)
GOOD_FONT = Font(name="Segoe UI", color="006100", size=10, bold=True)
THIN_BORDER = Border(bottom=Side(style="thin", color="D9D9D9"))
ALT_FILL = PatternFill("solid", fgColor="F2F7FB")

NUM_FMT_EH = '0.00000000'


def build_report(
    records: list[dict],
    image_map: dict[str, str],
    smiles_map: dict[str, str],
    xlsx_path: Path,
):
    """Build a polished Excel report with embedded molecule images.

    Parameters
    ----------
    records : list of parse_out_file() dicts
    image_map : {base_name: png_file_path}
    smiles_map : {base_name: SMILES}
    xlsx_path : destination .xlsx path
    """
    wb = Workbook()

    # ── Results sheet ─────────────────────────────────────
    ws = wb.active
    ws.title = "Results"

    headers = [
        "Structure",
        "Name",
        "SMILES",
        "Status",
        "Electronic Energy (Eh)",
        "Total Enthalpy (Eh)",
        "Entropy Correction (Eh)",
        "Final Gibbs Free Energy (Eh)",
        "Total Run Time",
        "Errors",
    ]
    ws.append(headers)

    # Header styling
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 36
    ws.freeze_panes = "A2"

    # Column widths
    col_widths = {
        "A": 22, "B": 18, "C": 28, "D": 12,
        "E": 24, "F": 22, "G": 24, "H": 28,
        "I": 20, "J": 42,
    }
    for letter, w in col_widths.items():
        ws.column_dimensions[letter].width = w

    # Data rows
    for i, r in enumerate(records):
        row = i + 2
        base = r["file"].replace(".out", "")
        smiles = smiles_map.get(base, "")
        err_str = "; ".join(r["errors"]) if r["errors"] else ""
        status = "OK" if r["normal_term"] else "FAILED"

        ws.cell(row=row, column=1, value="")  # image placeholder
        ws.cell(row=row, column=2, value=base)
        ws.cell(row=row, column=3, value=smiles)
        ws.cell(row=row, column=4, value=status)
        ws.cell(row=row, column=5, value=r.get("electronic_energy_eh"))
        ws.cell(row=row, column=6, value=r["enthalpy_eh"])
        ws.cell(row=row, column=7, value=r["entropy_eh"])
        ws.cell(row=row, column=8, value=r["gibbs_eh"])
        ws.cell(row=row, column=9, value=r["total_run_time"] or "")
        ws.cell(row=row, column=10, value=err_str)

        # Number format
        for col in [5, 6, 7, 8]:
            c = ws.cell(row=row, column=col)
            if c.value is not None:
                c.number_format = NUM_FMT_EH

        # Row height for image
        ws.row_dimensions[row].height = 85

        # Alternating row colour
        fill = ALT_FILL if i % 2 == 0 else PatternFill()
        for col in range(1, len(headers) + 1):
            c = ws.cell(row=row, column=col)
            c.font = NORMAL_FONT
            c.border = THIN_BORDER
            c.alignment = Alignment(vertical="center", wrap_text=True)
            if fill.fgColor and str(fill.fgColor.rgb) != "00000000":
                c.fill = fill

        # Status cell colour
        sc = ws.cell(row=row, column=4)
        if r["normal_term"]:
            sc.font = GOOD_FONT
        else:
            sc.font = ERROR_FONT
            ws.cell(row=row, column=10).font = ERROR_FONT

        # Embed molecule image
        png_path = image_map.get(base)
        if png_path and Path(png_path).exists():
            try:
                img = XlImage(png_path)
                img.width = 140
                img.height = 105
                ws.add_image(img, f"A{row}")
            except Exception:
                pass

    # ── Timings sheet ─────────────────────────────────────
    ws2 = wb.create_sheet("Timings")
    t_headers = ["File", "Module", "Time (sec)", "Percent (%)"]
    ws2.append(t_headers)
    for col in range(1, len(t_headers) + 1):
        c = ws2.cell(row=1, column=col)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")
    ws2.freeze_panes = "A2"

    for r in records:
        for t in r["timings"]:
            ws2.append([r["file"], t["module"], t["seconds"], t["pct"]])
            ri = ws2.max_row
            for col in range(1, len(t_headers) + 1):
                ws2.cell(row=ri, column=col).font = NORMAL_FONT
                ws2.cell(row=ri, column=col).border = THIN_BORDER
        if r["total_run_time"]:
            ws2.append([r["file"], "TOTAL RUN TIME", r["total_run_time"], ""])
            ri = ws2.max_row
            for col in range(1, len(t_headers) + 1):
                ws2.cell(row=ri, column=col).font = BOLD_FONT
                ws2.cell(row=ri, column=col).border = THIN_BORDER
        ws2.append([])

    for col_cells in ws2.columns:
        col_letter = get_column_letter(col_cells[0].column)
        lengths = [len(str(c.value)) for c in col_cells if c.value is not None]
        best = max(lengths) if lengths else 10
        ws2.column_dimensions[col_letter].width = min(max(best + 3, 10), 45)

    # ── Errors sheet (only when needed) ───────────────────
    if any(r["errors"] for r in records):
        ws3 = wb.create_sheet("Errors")
        e_headers = ["File", "Error Message"]
        ws3.append(e_headers)
        for col in range(1, len(e_headers) + 1):
            c = ws3.cell(row=1, column=col)
            c.font = HEADER_FONT
            c.fill = HEADER_FILL
            c.alignment = Alignment(horizontal="center")
        ws3.freeze_panes = "A2"

        for r in records:
            for e in r["errors"]:
                ws3.append([r["file"], e])
                ri = ws3.max_row
                for col in range(1, len(e_headers) + 1):
                    c = ws3.cell(row=ri, column=col)
                    c.fill = ERROR_FILL
                    c.font = ERROR_FONT
                    c.border = THIN_BORDER

        for col_cells in ws3.columns:
            col_letter = get_column_letter(col_cells[0].column)
            lengths = [len(str(c.value)) for c in col_cells if c.value is not None]
            best = max(lengths) if lengths else 10
            ws3.column_dimensions[col_letter].width = min(max(best + 3, 10), 80)

    wb.save(xlsx_path)
