"""
pipeline.py - Core pipeline: SMILES -> 3D -> ORCA input -> run -> parse -> report

Handles the full computational chemistry workflow:
  1. Convert SMILES to 3D geometries via RDKit force-field optimization
  2. Generate ORCA input files from settings + geometry
  3. Run ORCA calculations
  4. Parse .out files for thermodynamic data
  5. Build Excel report with embedded molecule structure images
"""

import io
import os
import re
import time
import shutil
import subprocess
from pathlib import Path
from datetime import datetime

from rdkit import Chem
from rdkit.Chem import AllChem, Draw
from PIL import Image as PILImage

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XlImage

ROOT = Path(__file__).resolve().parent

# ===================== USER SETTING =====================
ORCA_EXE = Path(r"C:\ORCA_6.1.1\orca.exe")
# ========================================================

INP_DIR = ROOT / "INP"
OUT_DIR = ROOT / "OUT"
MID_DIR = ROOT / "MID"


def make_run_stamp() -> str:
    """Readable folder name like 'Apr. 02 2026 10.37am'."""
    now = datetime.now()
    raw = now.strftime("%b. %d %Y %I.%M%p")
    return raw[:-2] + raw[-2:].lower()


# -------------------------------------------------------------------
#  SMILES -> 3D GEOMETRY + MOLECULE IMAGE
# -------------------------------------------------------------------

def validate_smiles(smiles: str) -> bool:
    """Return True if RDKit can parse the SMILES string."""
    return Chem.MolFromSmiles(smiles) is not None


def smiles_to_xyz(smiles: str):
    """Convert a SMILES string to optimized 3D coordinates and a 2D image.

    Returns:
        (xyz_block, png_bytes, num_atoms)
    Raises:
        ValueError on invalid SMILES or embedding failure.
    """
    mol = Chem.MolFromSmiles(smiles)
    if mol is None:
        raise ValueError(f"Invalid SMILES: {smiles}")

    mol_3d = Chem.AddHs(mol)

    # Embed 3D coordinates (ETKDGv3 with deterministic seed)
    params = AllChem.ETKDGv3()
    params.randomSeed = 42
    if AllChem.EmbedMolecule(mol_3d, params) == -1:
        # Fallback to basic embedding
        if AllChem.EmbedMolecule(mol_3d, randomSeed=42) == -1:
            raise ValueError(f"3D embedding failed for: {smiles}")

    # Force-field optimization: MMFF preferred, UFF fallback
    try:
        AllChem.MMFFOptimizeMolecule(mol_3d, maxIters=2000)
    except Exception:
        try:
            AllChem.UFFOptimizeMolecule(mol_3d, maxIters=2000)
        except Exception:
            pass  # proceed with un-optimised geometry

    # Build XYZ coordinate block
    conf = mol_3d.GetConformer()
    lines = []
    for i, atom in enumerate(mol_3d.GetAtoms()):
        pos = conf.GetAtomPosition(i)
        lines.append(
            f"  {atom.GetSymbol():2s}  {pos.x:14.8f}  {pos.y:14.8f}  {pos.z:14.8f}"
        )

    xyz_block = "\n".join(lines)
    num_atoms = mol_3d.GetNumAtoms()

    # Generate 2D structure image (PNG)
    mol_2d = Chem.MolFromSmiles(smiles)
    AllChem.Compute2DCoords(mol_2d)
    img = Draw.MolToImage(mol_2d, size=(400, 300))
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    png_bytes = buf.getvalue()

    return xyz_block, png_bytes, num_atoms


# -------------------------------------------------------------------
#  ORCA INPUT FILE GENERATION
# -------------------------------------------------------------------

def generate_inp(name: str, xyz_block: str, settings: dict) -> str:
    """Build ORCA .inp file content from XYZ coordinates and user settings."""
    functional = settings.get("functional", "B3LYP")
    basis_set = settings.get("basis_set", "def2-SVP")
    calc_type = settings.get("calc_type", "OPT FREQ")
    charge = settings.get("charge", 0)
    multiplicity = settings.get("multiplicity", 1)
    ram_mb = settings.get("ram_mb", 4000)
    cpus = settings.get("cpus", 4)
    extra_keywords = settings.get("extra_keywords", "").strip()
    extra_blocks = settings.get("extra_blocks", "").strip()

    keyword_line = f"! {functional} {basis_set} {calc_type}"
    if extra_keywords:
        keyword_line += f" {extra_keywords}"

    parts = [
        keyword_line,
        "",
        f"%maxcore {ram_mb}",
        "%pal",
        f"  nprocs {cpus}",
        "end",
    ]

    if extra_blocks:
        parts += ["", extra_blocks]

    parts += [
        "",
        f"* xyz {charge} {multiplicity}",
        xyz_block,
        "*",
        "",
    ]

    return "\n".join(parts)


# -------------------------------------------------------------------
#  ORCA JOB RUNNER
# -------------------------------------------------------------------

def fmt_hhmmss(seconds: float) -> str:
    s = int(seconds)
    h, rem = divmod(s, 3600)
    m, sec = divmod(rem, 60)
    return f"{h:02d}:{m:02d}:{sec:02d}"


def run_orca_job(
    base: str,
    inp_src: Path,
    mid_dir: Path,
    out_dir: Path,
    env: dict,
    status_dict: dict | None = None,
    pipeline_start: float | None = None,
) -> int:
    """Run a single ORCA job.  Returns process exit code (0 = success).

    If *status_dict* and *pipeline_start* are provided the function
    periodically updates status_dict["elapsed"] so the web UI can
    show a live timer.
    """
    job_dir = mid_dir / base
    job_out_dir = out_dir / base
    out_file = job_out_dir / f"{base}.out"
    err_file = job_out_dir / f"{base}.out.err"

    job_dir.mkdir(parents=True, exist_ok=True)
    job_out_dir.mkdir(parents=True, exist_ok=True)

    dst_inp = job_dir / f"{base}.inp"
    shutil.copy2(inp_src, dst_inp)

    tmpdir = job_dir / "_tmp"
    tmpdir.mkdir(parents=True, exist_ok=True)

    env_local = dict(env)
    env_local["TMPDIR"] = str(tmpdir)
    env_local["TEMP"] = str(tmpdir)
    env_local["TMP"] = str(tmpdir)

    with out_file.open("wb") as fout, err_file.open("wb") as ferr:
        try:
            p = subprocess.Popen(
                [str(ORCA_EXE), dst_inp.name],
                cwd=str(job_dir),
                stdout=fout,
                stderr=ferr,
                env=env_local,
                shell=False,
            )
        except Exception:
            return 1

        # Poll so we can update the live timer
        while p.poll() is None:
            if status_dict is not None and pipeline_start is not None:
                status_dict["elapsed"] = fmt_hhmmss(time.time() - pipeline_start)
            time.sleep(2)

        rc = p.returncode

    # Append stderr into .out, then delete .err
    if err_file.exists():
        try:
            with err_file.open("rb") as ef, out_file.open("ab") as of:
                of.write(b"\n")
                of.write(ef.read())
            err_file.unlink()
        except OSError:
            pass

    # Copy .xyz artefacts to the output subfolder
    for xyz in job_dir.glob("*.xyz"):
        try:
            shutil.copy2(xyz, job_out_dir / xyz.name)
        except Exception:
            pass

    return rc


# -------------------------------------------------------------------
#  ORCA OUTPUT PARSER
# -------------------------------------------------------------------

def parse_out_file(path: Path) -> dict:
    """Extract thermodynamic data, timings, and errors from an ORCA .out file."""
    text = path.read_text(errors="replace")

    result = {
        "file": path.name,
        "normal_term": False,
        "errors": [],
        "electronic_energy_eh": None,
        "enthalpy_eh": None,
        "entropy_eh": None,
        "entropy_kcal": None,
        "gibbs_eh": None,
        "timings": [],
        "total_run_time": None,
    }

    if "****ORCA TERMINATED NORMALLY****" in text:
        result["normal_term"] = True

    for line in text.splitlines():
        s = line.strip()
        if s.startswith("ORCA ABORT") or s.startswith("ERROR"):
            result["errors"].append(s)
        elif "ABORTING THE RUN" in s:
            result["errors"].append(s)

    if not result["normal_term"] and not result["errors"]:
        result["errors"].append("Job did not terminate normally (no specific error captured)")

    m = re.search(r"FINAL SINGLE POINT ENERGY\s+([-\d.]+)", text)
    if m:
        result["electronic_energy_eh"] = float(m.group(1))

    m = re.search(r"Total enthalpy\s+\.\.\.\s+([-\d.]+)\s*Eh", text)
    if m:
        result["enthalpy_eh"] = float(m.group(1))

    m = re.search(
        r"Total entropy correction\s+\.\.\.\s+([-\d.]+)\s*Eh\s+([-\d.]+)\s*kcal/mol",
        text,
    )
    if m:
        result["entropy_eh"] = float(m.group(1))
        result["entropy_kcal"] = float(m.group(2))

    m = re.search(r"Final Gibbs free energy\s+\.\.\.\s+([-\d.]+)\s*Eh", text)
    if m:
        result["gibbs_eh"] = float(m.group(1))

    m = re.search(r"TOTAL RUN TIME:\s*(.+)", text)
    if m:
        result["total_run_time"] = m.group(1).strip()

    timing_pat = re.compile(
        r"^([\w\s]+?)\s+\.{3}\s+([\d.]+)\s+sec\s+\(=\s+[\d.]+\s+min\)\s+([\d.]+)\s*%",
        re.MULTILINE,
    )
    for tm in timing_pat.finditer(text):
        result["timings"].append(
            {
                "module": tm.group(1).strip(),
                "seconds": float(tm.group(2)),
                "pct": float(tm.group(3)),
            }
        )

    return result


# -------------------------------------------------------------------
#  EXCEL REPORT WITH MOLECULE IMAGES
# -------------------------------------------------------------------

HEADER_FILL = PatternFill("solid", fgColor="1e3a5f")
HEADER_FONT = Font(name="Segoe UI", bold=True, color="FFFFFF", size=11)
NORMAL_FONT = Font(name="Segoe UI", size=10)
BOLD_FONT = Font(name="Segoe UI", size=10, bold=True)
ERROR_FILL = PatternFill("solid", fgColor="FFC7CE")
ERROR_FONT = Font(name="Segoe UI", color="9C0006", size=10)
GOOD_FONT = Font(name="Segoe UI", color="006100", size=10, bold=True)
THIN_BORDER = Border(bottom=Side(style="thin", color="D9D9D9"))
ALT_FILL = PatternFill("solid", fgColor="F2F7FB")

NUM_FMT_EH = '0.00000000'


def build_report(
    records: list[dict],
    image_map: dict[str, str],
    smiles_map: dict[str, str],
    xlsx_path: Path,
):
    """Build a polished Excel report with embedded molecule images.

    Parameters
    ----------
    records : list of parse_out_file() dicts
    image_map : {base_name: png_file_path}
    smiles_map : {base_name: SMILES}
    xlsx_path : destination .xlsx path
    """
    wb = Workbook()

    # ── Results sheet ─────────────────────────────────────
    ws = wb.active
    ws.title = "Results"

    headers = [
        "Structure",
        "Name",
        "SMILES",
        "Status",
        "Electronic Energy (Eh)",
        "Total Enthalpy (Eh)",
        "Entropy Correction (Eh)",
        "Final Gibbs Free Energy (Eh)",
        "Total Run Time",
        "Errors",
    ]
    ws.append(headers)

    # Header styling
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 36
    ws.freeze_panes = "A2"

    # Column widths
    col_widths = {
        "A": 22, "B": 18, "C": 28, "D": 12,
        "E": 24, "F": 22, "G": 24, "H": 28,
        "I": 20, "J": 42,
    }
    for letter, w in col_widths.items():
        ws.column_dimensions[letter].width = w

    # Data rows
    for i, r in enumerate(records):
        row = i + 2
        base = r["file"].replace(".out", "")
        smiles = smiles_map.get(base, "")
        err_str = "; ".join(r["errors"]) if r["errors"] else ""
        status = "OK" if r["normal_term"] else "FAILED"

        ws.cell(row=row, column=1, value="")  # image placeholder
        ws.cell(row=row, column=2, value=base)
        ws.cell(row=row, column=3, value=smiles)
        ws.cell(row=row, column=4, value=status)
        ws.cell(row=row, column=5, value=r.get("electronic_energy_eh"))
        ws.cell(row=row, column=6, value=r["enthalpy_eh"])
        ws.cell(row=row, column=7, value=r["entropy_eh"])
        ws.cell(row=row, column=8, value=r["gibbs_eh"])
        ws.cell(row=row, column=9, value=r["total_run_time"] or "")
        ws.cell(row=row, column=10, value=err_str)

        # Number format
        for col in [5, 6, 7, 8]:
            c = ws.cell(row=row, column=col)
            if c.value is not None:
                c.number_format = NUM_FMT_EH

        # Row height for image
        ws.row_dimensions[row].height = 85

        # Alternating row colour
        fill = ALT_FILL if i % 2 == 0 else PatternFill()
        for col in range(1, len(headers) + 1):
            c = ws.cell(row=row, column=col)
            c.font = NORMAL_FONT
            c.border = THIN_BORDER
            c.alignment = Alignment(vertical="center", wrap_text=True)
            if fill.fgColor and str(fill.fgColor.rgb) != "00000000":
                c.fill = fill

        # Status cell colour
        sc = ws.cell(row=row, column=4)
        if r["normal_term"]:
            sc.font = GOOD_FONT
        else:
            sc.font = ERROR_FONT
            ws.cell(row=row, column=10).font = ERROR_FONT

        # Embed molecule image
        png_path = image_map.get(base)
        if png_path and Path(png_path).exists():
            try:
                img = XlImage(png_path)
                img.width = 140
                img.height = 105
                ws.add_image(img, f"A{row}")
            except Exception:
                pass

    # ── Timings sheet ─────────────────────────────────────
    ws2 = wb.create_sheet("Timings")
    t_headers = ["File", "Module", "Time (sec)", "Percent (%)"]
    ws2.append(t_headers)
    for col in range(1, len(t_headers) + 1):
        c = ws2.cell(row=1, column=col)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")
    ws2.freeze_panes = "A2"

    for r in records:
        for t in r["timings"]:
            ws2.append([r["file"], t["module"], t["seconds"], t["pct"]])
            ri = ws2.max_row
            for col in range(1, len(t_headers) + 1):
                ws2.cell(row=ri, column=col).font = NORMAL_FONT
                ws2.cell(row=ri, column=col).border = THIN_BORDER
        if r["total_run_time"]:
            ws2.append([r["file"], "TOTAL RUN TIME", r["total_run_time"], ""])
            ri = ws2.max_row
            for col in range(1, len(t_headers) + 1):
                ws2.cell(row=ri, column=col).font = BOLD_FONT
                ws2.cell(row=ri, column=col).border = THIN_BORDER
        ws2.append([])

    for col_cells in ws2.columns:
        col_letter = get_column_letter(col_cells[0].column)
        lengths = [len(str(c.value)) for c in col_cells if c.value is not None]
        best = max(lengths) if lengths else 10
        ws2.column_dimensions[col_letter].width = min(max(best + 3, 10), 45)

    # ── Errors sheet (only when needed) ───────────────────
    if any(r["errors"] for r in records):
        ws3 = wb.create_sheet("Errors")
        e_headers = ["File", "Error Message"]
        ws3.append(e_headers)
        for col in range(1, len(e_headers) + 1):
            c = ws3.cell(row=1, column=col)
            c.font = HEADER_FONT
            c.fill = HEADER_FILL
            c.alignment = Alignment(horizontal="center")
        ws3.freeze_panes = "A2"

        for r in records:
            for e in r["errors"]:
                ws3.append([r["file"], e])
                ri = ws3.max_row
                for col in range(1, len(e_headers) + 1):
                    c = ws3.cell(row=ri, column=col)
                    c.fill = ERROR_FILL
                    c.font = ERROR_FONT
                    c.border = THIN_BORDER

        for col_cells in ws3.columns:
            col_letter = get_column_letter(col_cells[0].column)
            lengths = [len(str(c.value)) for c in col_cells if c.value is not None]
            best = max(lengths) if lengths else 10
            ws3.column_dimensions[col_letter].width = min(max(best + 3, 10), 80)

    wb.save(xlsx_path)
